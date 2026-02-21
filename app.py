from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash, Response
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import random
import string
from openpyxl import load_workbook
from flask_socketio import SocketIO, emit
from functools import wraps
import json
import os
from oauth2client.service_account import ServiceAccountCredentials

app = Flask(__name__)
app.secret_key = "attendance_secret_v1"
socketio = SocketIO(app, async_mode="threading")

# ================= AUTHENTICATION ================= #

def check_auth(username, password):
    return username == 'radmin' and password == 'radmin@radmin'

def authenticate():
    return Response(
        'Could not verify your access level for that URL.\n'
        'You have to login with proper credentials', 401,
        {'WWW-Authenticate': 'Basic realm="Login Required"'})

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return authenticate()
        return f(*args, **kwargs)
    return decorated

# ================= GOOGLE SHEETS ================= #

scope = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive"
]



service_account_info = json.loads(
    os.environ["GOOGLE_CREDENTIALS"]
)

creds = ServiceAccountCredentials.from_json_keyfile_dict(
    service_account_info, scope)

client = gspread.authorize(creds)
sheet = client.open("Students").sheet1

students_cache = {
    str(student["Enrollment"]): student
    for student in sheet.get_all_records()
}

# ================= GLOBAL VARIABLES ================= #

ADMIN_CODE = None
LECTURE_TOPIC = ""
LECTURE_DATE = ""
attendance_records = []

# ================= ROUTES ================= #

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/enroll")
def enroll():
    return render_template("enroll.html")


def fix_drive_url(url):
    """Converts a Google Drive sharing link to a direct image link."""
    if not isinstance(url, str):
        return url
    if "drive.google.com" in url and "id=" in url:
        file_id = url.split("id=")[1].split("&")[0]
        return f"https://lh3.googleusercontent.com/d/{file_id}"
    if "/file/d/" in url:
        file_id = url.split("/file/d/")[1].split("/")[0]
        return f"https://lh3.googleusercontent.com/d/{file_id}"
    return url

@app.route("/validate", methods=["POST"])
def validate():
    enrollment = request.form["enrollment"]
    latitude = request.form["latitude"]
    longitude = request.form["longitude"]

    if not latitude or not longitude:
        flash("Location not detected. Please enable location.")
        return redirect(url_for("enroll"))

    # 🔄 Always fetch fresh data (FAST ENOUGH)
    records = sheet.get_all_records()

    students_cache = {
        str(student["Enrollment"]): student
        for student in records
    }

    student = students_cache.get(enrollment)

    if not student:
        flash("Enrollment Number Not Found!")
        return redirect(url_for("enroll"))

    # Fix Image URL if it's a Drive link
    student["ImageURL"] = fix_drive_url(student.get("ImageURL", ""))

    return render_template(
        "verify.html",
        student=student,
        latitude=latitude,
        longitude=longitude
    )


@app.route("/submit_code", methods=["POST"])
def submit_code():
    code = request.form["code"]

    if code != ADMIN_CODE:
        flash("Invalid Admin Code!")
        # Re-render verify page so student doesn't have to restart enrollment
        mock_student = {
            "Enrollment": request.form["enrollment"],
            "Name": request.form["name"],
            "Section": request.form["section"],
            "Course": request.form["course"],
            "ImageURL": fix_drive_url(request.form["image_url"])
        }
        return render_template(
            "verify.html",
            student=mock_student,
            latitude=request.form["latitude"],
            longitude=request.form["longitude"]
        )

    return redirect(url_for(
        "lecture",
        enrollment=request.form["enrollment"],
        name=request.form["name"],
        latitude=request.form["latitude"],
        longitude=request.form["longitude"],
        section=request.form["section"],
        course=request.form["course"]
    ))

@app.route("/mark_attendance", methods=["POST"])
def mark_attendance():
    enrollment = request.form["enrollment"]

    if any(r["Enrollment"] == enrollment for r in attendance_records):
        return jsonify({"status": "already"})

    new_record = {
    "Enrollment": enrollment,
    "Name": request.form["name"],
    "Latitude": request.form["latitude"],
    "Longitude": request.form["longitude"],
    "Section": request.form["section"],
    "Course": request.form.get("course", "")
}

    attendance_records.append(new_record)
    attendance_records.sort(key=lambda x: int(x["Enrollment"]))

    socketio.emit("new_attendance", attendance_records)

    return jsonify({"status": "success"})


@app.route("/get_attendance")
def get_attendance():
    sorted_records = sorted(
        attendance_records,
        key=lambda x: int(x["Enrollment"])
    )
    return jsonify(sorted_records)


@app.route("/admin", methods=["GET", "POST"])
@requires_auth
def admin():
    global ADMIN_CODE, LECTURE_TOPIC, LECTURE_DATE

    if request.method == "POST" and not ADMIN_CODE:
        LECTURE_TOPIC = request.form["topic"]
        LECTURE_DATE = request.form["date"]
        ADMIN_CODE = str(random.randint(100000, 999999))

    return render_template(
        "admin.html",
        code=ADMIN_CODE,
        topic=LECTURE_TOPIC,
        date=LECTURE_DATE,
        total=len(attendance_records)
    )


@app.route("/invalidate")
@requires_auth
def invalidate():
    global ADMIN_CODE, LECTURE_TOPIC, LECTURE_DATE, attendance_records
    ADMIN_CODE = None
    LECTURE_TOPIC = ""
    LECTURE_DATE = ""
    attendance_records = []
    return redirect(url_for("admin"))


@app.route("/download")
@requires_auth
def download():
    if not attendance_records:
        return "No Attendance Data"

    filename = f"{LECTURE_DATE}_{LECTURE_TOPIC}.xlsx"

    df = pd.DataFrame(attendance_records)
    df.to_excel("attendance.xlsx", index=False)

    wb = load_workbook("attendance.xlsx")
    ws = wb.active

    ws.insert_rows(1, amount=3)

    ws["A1"] = f"Date: {LECTURE_DATE}"
    ws["A2"] = f"Topic: {LECTURE_TOPIC}"
    ws["A3"] = f"Total Students Present: {len(attendance_records)}"

    wb.save("attendance.xlsx")

    return send_file(
        "attendance.xlsx",
        as_attachment=True,
        download_name=filename
    )


@app.route("/reset")
@requires_auth
def reset():
    global attendance_records
    attendance_records = []
    return redirect(url_for("admin"))


@app.route("/lecture")
def lecture():
    return render_template(
        "lecture.html",
        topic=LECTURE_TOPIC,
        date=LECTURE_DATE
    )


if __name__ == "__main__":
    socketio.run(app, host="0.0.0.0", port=7860)